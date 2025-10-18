#!/usr/bin/env python3
"""Batch import multiple CFTS Excel files into database."""
import pandas as pd
import openpyxl
import json
import sys
import os
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple

from app.db.database import engine, SessionLocal, Base
from app.db.crud import bulk_create_cfts_requirements
from app.models.requirement import CFTSRequirement
from app.models.cfts_db import CFTSRequirementDB


class BatchImporter:
    """Batch import CFTS Excel files."""

    def __init__(self, excel_folder: str, md_scope_filter: bool = True):
        """
        Initialize batch importer.

        Args:
            excel_folder: Path to folder containing Excel files
            md_scope_filter: If True, only import records with MD Scope = "Yes"
        """
        self.excel_folder = Path(excel_folder)
        self.md_scope_filter = md_scope_filter
        self.report = {
            'total_files': 0,
            'success_files': [],
            'failed_files': [],
            'total_records': 0,
            'inserted_records': 0,
            'skipped_records': 0,
            'errors': []
        }

    def find_excel_files(self) -> List[Path]:
        """Find all Excel files in the folder."""
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(self.excel_folder.glob(ext))
        return sorted(excel_files)

    def extract_cfts_from_filename(self, filename: str) -> str:
        """
        Extract CFTS number from filename.

        Example: SYS1_CFTS091_Player Component Functions_SR26.xlsx -> CFTS091
        """
        match = re.search(r'CFTS\d+', filename)
        if match:
            return match.group(0)
        return ''

    def extract_cfts_name_from_filename(self, filename: str) -> str:
        """
        Extract CFTS name from filename.

        Example: SYS1_CFTS016_Anti-Theft_SR26.xlsx -> Anti-Theft
        Example: SYS1_CFTS091_Player Component Functions_SR26.xlsx -> Player Component Functions
        Example: SYS1_CFTS086 Manual and Automatic Emergency Call_SR26.xlsx -> Manual and Automatic Emergency Call
        """
        # Remove file extension
        name_without_ext = filename.replace('.xlsx', '').replace('.xls', '')

        # Try to match pattern: anything after CFTS\d+ and before _SR\d+
        # This handles both underscore and space separators after CFTS
        match = re.search(r'CFTS\d+[_ ](.+?)_SR\d+', name_without_ext)
        if match:
            cfts_name = match.group(1).strip()
            return cfts_name

        return ''

    def parse_excel_file(self, file_path: Path) -> Tuple[List[Dict], int, int]:
        """
        Parse a single Excel file.

        Returns:
            Tuple of (parsed_data, total_count, filtered_count)
        """
        try:
            # Extract CFTS number and name from filename
            cfts_number = self.extract_cfts_from_filename(file_path.name)
            if not cfts_number:
                raise Exception(f"Could not extract CFTS number from filename: {file_path.name}")

            cfts_name = self.extract_cfts_name_from_filename(file_path.name)

            # Read Excel file with pandas for data
            df = pd.read_excel(file_path)
            total_count = len(df)

            # Read Excel file with openpyxl for hyperlinks
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Extract hyperlinks from column A (ID column)
            hyperlinks = {}
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                cell = ws[f'A{row_idx}']
                if cell.hyperlink and cell.value:
                    polarian_id = str(cell.value).strip()
                    hyperlinks[polarian_id] = cell.hyperlink.target

            # Filter by Scope (MD Scope) if enabled
            if self.md_scope_filter:
                df = df[df['Scope'].str.strip().str.upper() == 'YES']

            filtered_count = len(df)

            # Convert to list of dicts
            data = []
            for _, row in df.iterrows():
                # Get Polarian ID from A column (ID)
                polarian_id = str(row.get('ID', '')).strip()  # A 欄: NEWR1L-xxxxx

                # Get ReqIF.ForeignID from C column
                reqif_id = row.get('ReqIF.ForeignID', '')
                req_id = str(reqif_id).strip() if pd.notna(reqif_id) else ''

                # Get hyperlink for this Polarian ID
                polarian_url = hyperlinks.get(polarian_id, '')

                # Use CFTS number and name from filename for classification
                record = {
                    'cfts_id': cfts_number,  # Use CFTS from filename
                    'cfts_name': cfts_name,  # CFTS name extracted from filename
                    'req_id': req_id,  # C 欄: ReqIF.ForeignID (數字，可能重複)
                    'polarian_id': polarian_id,  # A 欄: ID (NEWR1L-xxxxx)
                    'polarian_url': polarian_url,  # A 欄: Polarion hyperlink
                    'description': str(row.get('Description', '')).strip() if pd.notna(row.get('Description')) else '',
                    'spec_object_type': str(row.get('Spec Object Type', '')).strip() if pd.notna(row.get('Spec Object Type')) else ''
                }

                # Skip empty records (polarian_id is required as unique key)
                if record['polarian_id']:
                    data.append(record)

            return data, total_count, filtered_count

        except Exception as e:
            raise Exception(f"Error parsing {file_path.name}: {str(e)}")

    def import_to_database(self, data: List[Dict]) -> int:
        """
        Import data to database.

        Returns:
            Number of records successfully inserted
        """
        if not data:
            return 0

        # Convert to Pydantic models
        requirements = [CFTSRequirement(**item) for item in data]

        # Create database session
        db = SessionLocal()
        try:
            # Bulk insert (with duplicate handling)
            count = bulk_create_cfts_requirements(db, requirements)
            return count
        finally:
            db.close()

    def process_all_files(self) -> Dict:
        """Process all Excel files in the folder."""
        # Ensure database tables exist
        Base.metadata.create_all(bind=engine)

        # Find all Excel files
        excel_files = self.find_excel_files()
        self.report['total_files'] = len(excel_files)

        if not excel_files:
            print(f"No Excel files found in {self.excel_folder}")
            return self.report

        print(f"Found {len(excel_files)} Excel files")
        print(f"Scope filter: {'Enabled (Scope=Yes only)' if self.md_scope_filter else 'Disabled'}")
        print("-" * 80)

        # Process each file
        for idx, file_path in enumerate(excel_files, 1):
            cfts_num = self.extract_cfts_from_filename(file_path.name)
            print(f"\n[{idx}/{len(excel_files)}] Processing: {file_path.name}")
            print(f"  CFTS Number: {cfts_num}")

            try:
                # Parse Excel file
                data, total_count, filtered_count = self.parse_excel_file(file_path)
                print(f"  Total records: {total_count}")
                if self.md_scope_filter:
                    print(f"  Filtered (Scope=Yes): {filtered_count}")
                print(f"  Valid records: {len(data)}")

                # Import to database
                inserted_count = self.import_to_database(data)
                print(f"  Inserted: {inserted_count}")

                # Update report
                self.report['success_files'].append(file_path.name)
                self.report['total_records'] += total_count
                self.report['inserted_records'] += inserted_count
                self.report['skipped_records'] += (len(data) - inserted_count)

            except Exception as e:
                error_msg = str(e)
                print(f"  ERROR: {error_msg}")
                self.report['failed_files'].append(file_path.name)
                self.report['errors'].append({
                    'file': file_path.name,
                    'error': error_msg
                })

        return self.report

    def print_summary(self):
        """Print import summary report."""
        print("\n" + "=" * 80)
        print("IMPORT SUMMARY")
        print("=" * 80)
        print(f"Total files processed: {self.report['total_files']}")
        print(f"Successful: {len(self.report['success_files'])}")
        print(f"Failed: {len(self.report['failed_files'])}")
        print(f"\nTotal records read: {self.report['total_records']}")
        print(f"Successfully inserted: {self.report['inserted_records']}")
        print(f"Skipped (duplicates): {self.report['skipped_records']}")

        # Verify database
        db = SessionLocal()
        try:
            total_in_db = db.query(CFTSRequirementDB).count()
            print(f"\nTotal records in database: {total_in_db}")
        finally:
            db.close()

        if self.report['failed_files']:
            print("\nFailed files:")
            for file in self.report['failed_files']:
                print(f"  - {file}")

        if self.report['errors']:
            print("\nErrors:")
            for err in self.report['errors']:
                print(f"  - {err['file']}: {err['error']}")

        print("=" * 80)

        # Save report to file
        report_file = f"import_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(self.report, f, indent=2, ensure_ascii=False)
        print(f"\nDetailed report saved to: {report_file}")


def main():
    """Main function."""
    if len(sys.argv) < 2:
        print("Usage: python batch_import_excel.py <excel_folder> [--no-filter]")
        print("\nOptions:")
        print("  --no-filter    Import all records (don't filter by Scope)")
        sys.exit(1)

    excel_folder = sys.argv[1]
    md_scope_filter = '--no-filter' not in sys.argv

    if not os.path.isdir(excel_folder):
        print(f"Error: {excel_folder} is not a valid directory")
        sys.exit(1)

    # Create importer and process files
    importer = BatchImporter(excel_folder, md_scope_filter)
    importer.process_all_files()
    importer.print_summary()


if __name__ == "__main__":
    main()
