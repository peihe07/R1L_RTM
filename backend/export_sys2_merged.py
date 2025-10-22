#!/usr/bin/env python3
"""
Export all SYS2 requirements from database to a single merged Excel file.
"""
import sys
from pathlib import Path
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add app directory to path
sys.path.insert(0, str(Path(__file__).parent))

from app.models.sys2_requirement import SYS2RequirementDB
from app.db.database import DATABASE_URL
import os


def export_merged_sys2():
    """Export all SYS2 requirements to a single Excel file."""

    print("=" * 80)
    print("匯出合併 SYS2 要件到 Excel 檔案")
    print("=" * 80)

    # 連接資料庫
    database_url = DATABASE_URL
    print(f"\n連接資料庫: {database_url}")

    engine = create_engine(database_url)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        # 查詢所有 SYS2 要件，依 CFTS ID 和 Melco ID 排序
        print("\n查詢資料庫中的 SYS2 要件...")
        requirements = db.query(SYS2RequirementDB).order_by(
            SYS2RequirementDB.cfts_id,
            SYS2RequirementDB.melco_id
        ).all()

        if not requirements:
            print("❌ 資料庫中沒有 SYS2 要件資料")
            return

        print(f"✅ 找到 {len(requirements)} 筆 SYS2 要件")

        # 統計每個 CFTS 的要件數量
        cfts_stats = {}
        for req in requirements:
            cfts_id = req.cfts_id
            cfts_stats[cfts_id] = cfts_stats.get(cfts_id, 0) + 1

        print(f"\n包含 {len(cfts_stats)} 個 CFTS:")
        for cfts_id in sorted(cfts_stats.keys()):
            print(f"  {cfts_id}: {cfts_stats[cfts_id]} 筆")

        # 轉換為 DataFrame
        print("\n轉換資料為 Excel 格式...")
        data = []
        for req in requirements:
            data.append({
                'CFTS ID': req.cfts_id,
                'CFTS Name': req.cfts_name,
                'Melco ID (要件ID)': req.melco_id,
                '要件(英語)': req.requirement_en,
                '理由(英語)': req.reason_en,
                '補足(英語)': req.supplement_en,
                '種別': req.type,
                '関連要件ID': req.related_requirement_ids,
                '(R1L_SR21CFTS)': req.r1l_sr21cfts,
                '(R1L_SR22CFTS)': req.r1l_sr22cfts,
                '(R1L_SR23CFTS)': req.r1l_sr23cfts,
                '(R1L_SR24CFTS)': req.r1l_sr24cfts,
                '確認フェーズ': req.confirmation_phase,
                '検証基準': req.verification_criteria,
            })

        df = pd.DataFrame(data)

        # 匯出到 Excel
        output_path = Path('/data/SYS2/R1L_SYS2_ALL_MERGED.xlsx')
        print(f"\n匯出到: {output_path}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All SYS2 Requirements', index=False)

            # 取得 workbook 和 worksheet
            workbook = writer.book
            worksheet = writer.sheets['All SYS2 Requirements']

            # 設定樣式
            print("設定 Excel 樣式...")

            # 標題列樣式
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 設定欄寬
            column_widths = {
                'A': 12,  # CFTS ID
                'B': 30,  # CFTS Name
                'C': 20,  # Melco ID
                'D': 50,  # 要件(英語)
                'E': 40,  # 理由(英語)
                'F': 40,  # 補足(英語)
                'G': 15,  # 種別
                'H': 30,  # 関連要件ID
                'I': 30,  # R1L_SR21CFTS
                'J': 30,  # R1L_SR22CFTS
                'K': 30,  # R1L_SR23CFTS
                'L': 30,  # R1L_SR24CFTS
                'M': 15,  # 確認フェーズ
                'N': 40,  # 検証基準
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 為不同 CFTS 設定不同的背景顏色
            cfts_colors = {
                'CFTS004': 'E8F4F8',
                'CFTS009': 'FFF4E6',
                'CFTS010': 'E8F5E9',
                'CFTS011': 'FFF3E0',
                'CFTS012': 'F3E5F5',
                'CFTS014': 'E1F5FE',
                'CFTS015': 'FFF9C4',
                'CFTS016': 'FCE4EC',
                'CFTS019': 'E0F2F1',
                'CFTS020': 'F1F8E9',
                'CFTS021': 'E3F2FD',
                'CFTS022': 'FBE9E7',
                'CFTS025': 'F9FBE7',
                'CFTS026': 'FFE0B2',
                'CFTS032': 'F8BBD0',
                'CFTS041': 'DCEDC8',
                'CFTS042': 'FFCCBC',
                'CFTS044': 'CFD8DC',
            }

            # 預設顏色（淺灰）
            default_color = 'F5F5F5'

            # 套用資料列樣式
            cell_alignment = Alignment(vertical='top', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            current_cfts = None
            current_color = None

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                cfts_id = worksheet[f'A{row_idx}'].value

                # 取得該 CFTS 的顏色
                if cfts_id != current_cfts:
                    current_cfts = cfts_id
                    current_color = cfts_colors.get(cfts_id, default_color)

                fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = fill

            # 凍結首列和前三欄（CFTS ID, CFTS Name, Melco ID）
            worksheet.freeze_panes = 'D2'

            # 啟用自動篩選
            worksheet.auto_filter.ref = worksheet.dimensions

        print(f"\n✅ 成功匯出 {len(df)} 筆要件到: {output_path}")
        print(f"\n檔案大小: {output_path.stat().st_size / 1024 / 1024:.2f} MB")

        # 顯示摘要統計
        print("\n" + "=" * 80)
        print("匯出摘要")
        print("=" * 80)
        print(f"總筆數: {len(df)}")
        print(f"CFTS 數量: {len(cfts_stats)}")
        print(f"\n各 CFTS 統計:")
        for cfts_id in sorted(cfts_stats.keys()):
            count = cfts_stats[cfts_id]
            percentage = count / len(df) * 100
            print(f"  {cfts_id}: {count:4d} 筆 ({percentage:5.1f}%)")

        # 分析跨 CFTS 參考
        print("\n" + "=" * 80)
        print("跨 CFTS 參考分析")
        print("=" * 80)

        cross_ref_count = 0
        for idx, row in df.iterrows():
            sr21 = str(row['(R1L_SR21CFTS)'])
            cfts_id = row['CFTS ID']

            # 檢查是否包含其他 CFTS 的參考
            import re
            cfts_matches = re.findall(r'CFTS(\d+)', sr21)

            for match in cfts_matches:
                other_cfts = f'CFTS{match}'
                if other_cfts != cfts_id:
                    cross_ref_count += 1

        print(f"總共發現 {cross_ref_count} 個跨 CFTS 參考")
        print("\n提示：")
        print("  - 可以使用 Excel 的篩選功能快速找到特定 CFTS 的要件")
        print("  - 使用 Ctrl+F 搜尋 'CFTS021-822' 等參考編號")
        print("  - 不同 CFTS 用不同顏色標示，方便識別")
        print("  - 前三欄（CFTS ID, CFTS Name, Melco ID）已凍結，方便橫向捲動")

    except Exception as e:
        print(f"\n❌ 匯出失敗: {e}")
        import traceback
        traceback.print_exc()

    finally:
        db.close()


if __name__ == "__main__":
    export_merged_sys2()
